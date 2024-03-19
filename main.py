
from __future__ import annotations

import hashlib
import hmac
import itertools
import os
import secrets
import typing as t
import unicodedata
import openpyxl
PBKDF2_ROUNDS = 2048


class ConfigurationError(Exception):
    pass



class Mnemonic(object):
    def __init__(self, language: str = "english", wordlist: list[str] | None = None):
        self.radix = 2048
        self.language = language

        if wordlist is None:
            d = os.path.join(os.path.dirname(__file__), f"wordlist.txt")
            if os.path.exists(d) and os.path.isfile(d):
                with open(d, "r", encoding="utf-8") as f:
                    wordlist = [w.strip() for w in f.readlines()]
            else:
                raise ConfigurationError("Language not detected")

        if len(wordlist) != self.radix:
            raise ConfigurationError(f"Wordlist must contain {self.radix} words.")

        self.wordlist = wordlist
        # Japanese must be joined by ideographic space
        self.delimiter = "\u3000" if language == "japanese" else " "

    def to_mnemonic(self, data: bytes) -> str:
        if len(data) not in [16, 20, 24, 28, 32]:
            raise ValueError(
                f"Data length should be one of the following: [16, 20, 24, 28, 32], but it is not {len(data)}."
            )
        h = hashlib.sha256(data).hexdigest()
        b = (
            bin(int.from_bytes(data, byteorder="big"))[2:].zfill(len(data) * 8)
            + bin(int(h, 16))[2:].zfill(256)[: len(data) * 8 // 32]
        )
        result = []
        for i in range(len(b) // 11):
            idx = int(b[i * 11 : (i + 1) * 11], 2)
            result.append(self.wordlist[idx])
        return self.delimiter.join(result)

    def check(self, mnemonic: str) -> bool:
        mnemonic_list = self.normalize_string(mnemonic).split(" ")
        # list of valid mnemonic lengths
        if len(mnemonic_list) not in [12, 15, 18, 21, 24]:
            return False
        try:
            idx = map(
                lambda x: bin(self.wordlist.index(x))[2:].zfill(11), mnemonic_list
            )
            b = "".join(idx)
        except ValueError:
            return False
        l = len(b)  # noqa: E741
        d = b[: l // 33 * 32]
        h = b[-l // 33 :]
        nd = int(d, 2).to_bytes(l // 33 * 4, byteorder="big")
        nh = bin(int(hashlib.sha256(nd).hexdigest(), 16))[2:].zfill(256)[: l // 33]
        return h == nh

def savaMnemonics(mnemonics, fileName):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for i, mnemonic in enumerate(mnemonics, start=1):
        # for j, word in enumerate(mnemonic, start=1):
            sheet.cell(row=i, column=1).value = i
            sheet.cell(row=i, column=2).value = mnemonic
    workbook.save(fileName)

def calculateSha256(message):
    return hashlib.sha256((message).encode('utf-8')).hexdigest()

def main() -> None:
    prefix = ""
    numMnemonics = 1000
    mnemonics = []
    fileName = "mnemonics.xlsx"
    m = Mnemonic("english")
    for i in range(1, numMnemonics + 1):
        entropy = prefix + str(i)
        times = int(calculateSha256(entropy)[-1:], 16) + 1
        entropy = calculateSha256(entropy)
        data = bytes.fromhex(entropy)
        mnemonics.append(m.to_mnemonic(data))
    
    savaMnemonics(mnemonics, fileName)


if __name__ == "__main__":
    main()
